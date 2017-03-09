# setting alias to make it clearer and not confuse classes
import openpyxl as xl

# load workbook
wb = xl.load_workbook("C:\\Users\\u8008701\\Desktop\\Website NULL in TR while they exist.xlsx")

# get sheet names
# print(wb.get_sheet_names())

# choose sheet
sheet = wb.get_sheet_by_name("Sheet1")

# print(sheet.max_row)
# print(sheet.max_column)

# iterating through the rows in first columns, range(1,sheet.max_row) won't include the last row
# for x in range(1,sheet.max_row+1):
#     print(x)

# getting the value from a specific cell
# .value is required to get the value, otherwise we get the object, only applies to cells
print(sheet['A2'].value)
print(sheet.cell(row=2,column=1).value)

row_count = 0
for row in sheet.iter_rows():
    row_count +=1
    for cell in row:
        print(cell.value)
    if row_count == 5: break